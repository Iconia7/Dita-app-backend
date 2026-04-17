import re
from datetime import datetime
from decimal import Decimal

import openpyxl
import docx
from django.utils import timezone


def parse_date_string(date_str):
    if not date_str:
        return None
    match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", str(date_str))
    if match:
        day, month, year = match.groups()
        if len(year) == 2:
            year = "20" + year
        try:
            return datetime(int(year), int(month), int(day)).date()
        except ValueError:
            return None
    return None


def parse_time_range(time_str):
    """Returns (start_time_obj, end_time_obj, duration_decimal)"""
    if not time_str:
        return None, None, 0
    s = str(time_str).upper().replace(".", ":").replace(" ", "")
    try:
        parts = s.split("-")
        if len(parts) != 2:
            return None, None, 0

        def parse_t(t):
            for fmt in ("%I:%M%p", "%H:%M"):
                try:
                    return datetime.strptime(t, fmt)
                except ValueError:
                    continue
            return None

        start_dt = parse_t(parts[0])
        end_dt = parse_t(parts[1])

        if not start_dt or not end_dt:
            return None, None, 0

        diff = (end_dt - start_dt).total_seconds()
        if diff < 0:
            diff += 86400
        duration = Decimal(diff) / Decimal(3600)

        return start_dt.time(), end_dt.time(), duration
    except Exception as e:
        print(f"Error parsing time range: {e}")
        return None, None, 0


def expand_course_codes(val):
    """
    Expands shorthand course codes like BIL112X/Y into ['BIL112X', 'BIL112Y']
    and LLD345/407 into ['LLD345', 'LLD407'].

    Handles patterns such as:
        ACS211A/SIT211A           -> ['ACS211A', 'SIT211A']
        ACS213A/MIS221A/SIT213A   -> ['ACS213A', 'MIS221A', 'SIT213A']
        LLB 111/204/203A/419A     -> ['LLB111', 'LLB204', 'LLB203A', 'LLB419A']
        ACT 200/LLB212/307/422    -> ['ACT200', 'LLB212', 'LLB307', 'LLB422']
        LLB 213/315/407/214       -> ['LLB213', 'LLB315', 'LLB407', 'LLB214']
        BIL112X/Y                 -> ['BIL112X', 'BIL112Y']
    """
    if not val:
        return []

    parts = re.split(r"[/&\n,;]", str(val))
    expanded = []
    last_prefix = ""      # letter-only prefix of the last full code, e.g. "LLB"
    last_full_code = ""   # last full normalised code, e.g. "LLB111"

    for p in parts:
        # Normalise: remove ALL whitespace and uppercase
        p = re.sub(r"\s+", "", p).upper()
        if not p or p in ["CHAPEL", "BREAK", "NONE"]:
            continue

        # ── Pattern A: full course code  e.g. "ACS211A", "LLB111", "SIT213A"
        #    2+ leading letters, 1+ digits, optional trailing letters
        if re.match(r"^[A-Z]{2,}\d+[A-Z]*$", p):
            expanded.append(p)
            last_full_code = p
            m = re.match(r"^([A-Z]+)", p)
            last_prefix = m.group(1) if m else last_prefix

        # ── Pattern B: digits + optional trailing letters  e.g. "204", "203A", "419A"
        #    No leading letters — inherit last_prefix
        elif re.match(r"^\d+[A-Z]*$", p) and last_prefix:
            expanded.append(last_prefix + p)
            # Keep last_prefix unchanged for subsequent shorthand codes

        # ── Pattern C: single/double letter suffix  e.g. "Y" in "BIL112X/Y"
        #    Replace the last N chars of the previous full code
        elif re.match(r"^[A-Z]{1,2}$", p) and last_full_code and len(last_full_code) > len(p):
            expanded.append(last_full_code[: -len(p)] + p)

        # ── Fallback: keep as-is (unexpected format)
        else:
            if len(p) >= 3:
                expanded.append(p)
            last_full_code = p
            m = re.match(r"^([A-Z]+)", p)
            last_prefix = m.group(1) if m else last_prefix

    return expanded


def process_exam_excel(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    all_exams = []
    DAYS_SET = {"MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Normalize sheet name for venue prefix (clean up extra spaces)
        display_sheet_name = sheet_name.replace(" ", "").upper()

        rows = list(sheet.iter_rows(values_only=True))
        i = 0
        col_map = {}

        while i < len(rows):
            row = rows[i]
            if not row:
                i += 1
                continue

            row_text = " ".join([str(x).upper() for x in row if x])

            # 1. HEADER ROW DETECTION
            # Check if any cell in the row contains a day from our set
            is_header = any(day in row_text for day in DAYS_SET)

            if is_header:
                col_map = {}
                active_date = None
                for idx, cell in enumerate(row):
                    if not cell:
                        # Propagate active date across merged cells
                        if active_date:
                            col_map[idx] = {"date": active_date}
                        continue

                    new_date = parse_date_string(cell)
                    if new_date:
                        active_date = new_date
                        col_map[idx] = {"date": active_date}
                    elif active_date:
                        col_map[idx] = {"date": active_date}

                # 2. TIME ROW (Expected immediately below header)
                if i + 1 < len(rows):
                    i += 1  # Move to the time row
                    time_row = rows[i]
                    final_map = {}
                    for idx, meta in col_map.items():
                        if idx < len(time_row) and time_row[idx]:
                            s_time, e_time, dur = parse_time_range(time_row[idx])
                            if dur > 0:
                                meta["start"] = s_time
                                meta["end"] = e_time
                                meta["duration"] = dur
                                final_map[idx] = meta
                    col_map = final_map

            # 3. DATA ROW
            elif col_map and row[0]:
                venue_label = str(row[0]).strip()
                if venue_label.upper() not in ["ROOM", "NOTE:", "LUNCH", "BREAK", "CHAPEL"]:
                    # Prepend sheet name for unique campus venues
                    full_venue = f"{display_sheet_name} - {venue_label}"

                    for col_idx, meta in col_map.items():
                        if col_idx < len(row) and row[col_idx]:
                            val = str(row[col_idx]).strip()

                            # Use the new robust expansion logic
                            codes = expand_course_codes(val)

                            for code in codes:
                                if len(code) >= 3:
                                    naive_dt = datetime.combine(meta["date"], meta["start"])
                                    # Make the datetime aware of Africa/Nairobi offset
                                    full_datetime = timezone.make_aware(naive_dt)
                                    all_exams.append(
                                        {
                                            "course_code": code,
                                            "title": code,
                                            "date": full_datetime,
                                            "end_time": meta["end"],
                                            "venue": full_venue,
                                            "duration_hours": meta["duration"],
                                        }
                                    )
            i += 1
    return all_exams

def process_nursing_exam_docx(file_path):
    """
    Parses the Nursing School Word document exam timetable.
    Handles merged cells, campus labels, and multi-exam cells.
    """
    doc = docx.Document(file_path)
    all_exams = []

    current_date = None
    last_campus = ""

    for table in doc.tables:
        for row in table.rows:
            # Use raw cell access to resolve merged cells (python-docx repeats text for merges)
            cells = [c.text.strip() for c in row.cells]

            # Basic validations (The nursing table has ~11 columns)
            if not cells or len(cells) < 10:
                continue

            # 1. Date Propagation
            # The date usually appears in the first cell of a new day block.
            # We check this BEFORE the header skip to ensure we don't skip the day label.
            potential_date = parse_date_string(cells[0])
            if potential_date:
                current_date = potential_date

            # 2. Header Detection & Skip
            if "COORDINATOR" in cells[2].upper() or "HRS" in cells[4].upper():
                continue

            if not current_date:
                continue

            # 3. Campus Propagation
            campus = cells[1].strip() or last_campus
            if campus:
                last_campus = campus

            # 4. Helper for extracting exams from slots (Morning/Afternoon)
            def _extract_from_slot(course_text, hrs_text, venue_text, default_time_range):
                if not course_text or "NONE" in course_text.upper():
                    return

                # Time parsing
                s_time, e_time, dur_fallback = parse_time_range(default_time_range)
                if not s_time:
                    return

                # Split by newlines for cells containing multiple units
                units = [u.strip() for u in course_text.split('\n') if u.strip()]
                durations = [d.strip() for d in hrs_text.split('\n') if d.strip()]
                venues = [v.strip() for v in venue_text.split('\n') if v.strip()]

                for idx, unit_str in enumerate(units):
                    # Use Regex to hunt for codes like NUR 123 in the string
                    codes = re.findall(r'([A-Z]{2,}\s*\d+[A-Z]*)', unit_str.upper())
                    if not codes:
                        continue

                    # Match duration and venue per unit if available
                    d = durations[idx] if idx < len(durations) else (durations[0] if durations else str(dur_fallback))
                    v = venues[idx] if idx < len(venues) else (venues[0] if venues else "DMMLC")

                    # Add campus prefix as requested
                    full_venue = f"{campus} - {v}" if campus else v

                    for code_match in codes:
                        code = code_match.replace(' ', '') # Normalise NUR 123 to NUR123
                        naive_dt = datetime.combine(current_date, s_time)
                        full_datetime = timezone.make_aware(naive_dt)

                        all_exams.append({
                            "course_code": code,
                            "title": unit_str,
                            "date": full_datetime,
                            "end_time": e_time,
                            "venue": full_venue,
                            "duration_hours": Decimal(d) if d.replace('.', '', 1).isdigit() else dur_fallback,
                        })

            # Process Morning Slot (Cols 3, 4, 5)
            _extract_from_slot(cells[3], cells[4], cells[5], "8.30AM-11.30AM")

            # Process Afternoon Slot (Cols 7, 8, 9)
            _extract_from_slot(cells[7], cells[8], cells[9], "1.30PM-4.30PM")

    return all_exams
